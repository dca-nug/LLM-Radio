# ============================================================
# build_form_from_unik.py
# Bangun form anotasi dari TB_teks_unik.xlsx (ID_unik, teks, n_baris).
# ID + bacaan pre-filled; kolom label kosong + dropdown.
# Jalankan SETELAH ambil_unik.py.
# ============================================================
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from config import DATA_DIR

INPUT = DATA_DIR / 'TB_teks_unik.xlsx'          # ID_unik, teks, n_baris
OUT   = DATA_DIR / 'Form_Anotasi_TB_CXR.xlsx'

u = pd.read_excel(INPUT)
ids  = u['ID_unik'].astype(str).tolist()
teks = u['teks'].astype(str).tolist()
nbar = u['n_baris'].tolist() if 'n_baris' in u.columns else ['']*len(u)
N = len(u)

wb = Workbook(); ws = wb.active; ws.title = "Anotasi"
headers = ["ID_unik","Keterangan Rontgen","n_baris",
           "is_radiologi","imp_tb","infiltrat","nodul","kalsifikasi",
           "fibrotik","kavitas","kardiomegali","catatan"]
ws.append(headers)

hdr=PatternFill("solid",start_color="1F4E78"); gate=PatternFill("solid",start_color="FFF2CC")
imp=PatternFill("solid",start_color="DDEBF7"); fnd=PatternFill("solid",start_color="E2EFDA")
crd=PatternFill("solid",start_color="FCE4D6"); lock=PatternFill("solid",start_color="F2F2F2")
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

for col,h in enumerate(headers,1):
    c=ws.cell(1,col,h); c.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
    c.fill=hdr; c.border=bd
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

for r in range(N):
    row=r+2
    ws.cell(row,1,ids[r]); ws.cell(row,2,teks[r]); ws.cell(row,3,nbar[r])
    for col in range(1,13):
        cell=ws.cell(row,col); cell.font=Font(name="Arial",size=11); cell.border=bd
        cell.alignment=(Alignment(wrap_text=True,vertical="top") if col in (2,12)
                        else Alignment(horizontal="center",vertical="center"))
        if col in (1,2,3): cell.fill=lock
    ws.cell(row,4).fill=gate                       # is_radiologi
    ws.cell(row,5).fill=imp                         # imp_tb
    for col in range(6,11): ws.cell(row,col).fill=fnd
    ws.cell(row,11).fill=crd

dv_rad=DataValidation(type="list",formula1='"radiologi,artefak"',allow_blank=True,showErrorMessage=True)
dv_rad.errorTitle="Nilai tidak valid"; dv_rad.error="Pilih: radiologi / artefak"
ws.add_data_validation(dv_rad); dv_rad.add(f"D2:D{N+1}")

dv_imp=DataValidation(type="list",formula1='"aktif,suspek,inaktif,tidak"',allow_blank=True,showErrorMessage=True)
dv_imp.errorTitle="Nilai tidak valid"; dv_imp.error="Pilih: aktif / suspek / inaktif / tidak"
ws.add_data_validation(dv_imp); dv_imp.add(f"E2:E{N+1}")

dv_bin=DataValidation(type="list",formula1='"ada,tidak"',allow_blank=True,showErrorMessage=True)
dv_bin.errorTitle="Nilai tidak valid"; dv_bin.error="Pilih: ada / tidak"
ws.add_data_validation(dv_bin); dv_bin.add(f"F2:K{N+1}")

for col,w in zip("ABCDEFGHIJKL",[8,50,8,13,11,11,10,12,11,10,14,28]):
    ws.column_dimensions[col].width=w
ws.freeze_panes="A2"; ws.row_dimensions[1].height=42

wp=wb.create_sheet("Petunjuk"); wp.column_dimensions["A"].width=100
lines=[
 (f"PETUNJUK ANOTASI — Laporan Radiologi Toraks Skrining TB ({N} teks unik)",True),
 ("ID + bacaan SUDAH terisi (abu-abu) — jangan diubah. Isi HANYA kolom label (dropdown).",False),
 ("Satuan = TEKS. Nilai apa yang TERTULIS, bukan simpulan klinis Anda. Kosongkan bila ragu.",False),
 ("",False),
 ("is_radiologi (pilih dulu): radiologi / artefak",True),
 ("   artefak  = teks BUKAN bacaan foto (mis. hamil, menolak, tdl, status pengobatan, catatan admin).",False),
 ("   radiologi= teks adalah bacaan/kesan foto toraks.",False),
 ("   Bila 'artefak', kolom label lain boleh dikosongkan.",False),
 ("",False),
 ("imp_tb (pilih SATU; dasar = kalimat KESAN, bukan temuan): aktif / suspek / inaktif / tidak",True),
 ("   aktif=TB aktif tegas | suspek=ragu (suspek/curiga/DD) | inaktif=lama/bekas/post/scar | tidak=tak disebut/disingkirkan",False),
 ("",False),
 ("temuan (ada/tidak, independen): infiltrat, nodul, kalsifikasi, fibrotik, kavitas",True),
 ("   'ada' bila disebut ada; 'tidak' bila tak disebut ATAU dinegasi. Adanya temuan TIDAK otomatis TB.",False),
 ("kardiomegali (ada/tidak): 'ada' bila jantung membesar / CTR meningkat disebut.",True),
 ("",False),
 ("TIDAK dianotasi: 'normal' & 'non-TB abnormal' dihitung otomatis. n_baris = jumlah baris asli terwakili.",False),
 ("",False),
 ("INDEPENDENSI: tiap anotator pakai SALINAN SENDIRI. Jangan lihat anotasi lain / keluaran sistem.",True),
 ("Rujukan lengkap: Annotation_Protocol_TB_CXR (dokumen terpisah).",False),
]
for i,(t,b) in enumerate(lines,1):
    c=wp.cell(i,1,t); c.font=Font(name="Arial",bold=b,size=(13 if(b and i==1)else 11),
        color=("1F4E78" if b else "000000")); c.alignment=Alignment(wrap_text=True,vertical="top")

wb.save(OUT)
print(f'Form: {OUT} | {N} teks | kolom label: is_radiologi + imp_tb + 5 temuan + kardiomegali')
